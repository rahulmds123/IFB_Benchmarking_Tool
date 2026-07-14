import pandas as pd
import os, uuid, zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font
import io

os.makedirs("outputs", exist_ok=True)

# ---- JOB STORE ----
# Each entry: {"normalized": {company: df}, "product_type": "washing_machine"|"ac",
#              "analysis_spec_cols": [...]}
# analysis_spec_cols is stored per-job because washing machine and AC BOMs
# don't share the same spec columns (AC has no Thickness, has Colour instead)
# — every endpoint that inspects spec columns reads this instead of assuming
# a single global column set.
JOB_STORE = {}

ASSEMBLY_ALIASES = {
    "motor clutch assembly": "clutch-motor assembly",
}

SPEC_COLS = [
    "Manufacturing Process",
    "Position (Assembled where)",
    "Dimensions/Specs(mm)",
    "Thickness (mm)",
    "Weight (Grams/piece)",
    "Number of Part",
    "Total Weight",
    "Total Assembly weight",
    "Material",
    "Characteristic (eg-special point)",
]

KEY_COLS = {"Sr No", "Assembly Area", "No of Parts", "Component name"}

ANALYSIS_SPEC_COLS = [
    "Manufacturing Process",
    "Position (Assembled where)",
    "Dimensions/Specs(mm)",
    "Thickness (mm)",
    "Material",
     "Weight (Grams/piece)",
    "Characteristic (eg-special point)"
]


# ---------------- LOAD ----------------
def load_bm_sheet(file):
    xls = pd.ExcelFile(file)
    if "Detailed BM Chart" in xls.sheet_names:
        return pd.read_excel(xls, sheet_name="Detailed BM Chart")
    for sheet in xls.sheet_names:
        if "bm" in sheet.lower():
            return pd.read_excel(xls, sheet_name=sheet)
    raise ValueError("No BM sheet found")


# ---------------- CLEAN ----------------
def clean_df(df):
    df.columns = (
        df.columns
        .str.replace('\n', ' ', regex=False)
        .str.replace('.', '', regex=False)
        .str.replace(r'\s+', ' ', regex=True)
        .str.strip()
    )
    for col in df.select_dtypes(include="object").columns:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace('\n', ' ', regex=False)
            .str.replace(r'\s+', ' ', regex=True)
            .str.strip()
        )
    df = df.replace("nan", pd.NA).replace("NA", pd.NA)
    df["Assembly Area"] = (
        df["Assembly Area"]
        .astype(str).str.strip().str.lower()
        .str.replace(r'\s+', ' ', regex=True)
        .replace("nan", pd.NA)
        .replace(ASSEMBLY_ALIASES)
    )
    df["Component name"] = (
        df["Component name"]
        .astype(str).str.strip().str.lower()
        .replace("nan", pd.NA)
    )
    df["No of Parts"] = pd.to_numeric(df["No of Parts"], errors="coerce").round(2)
    df["Assembly Area"] = df["Assembly Area"].ffill()
    return df


# ---------------- MASTER BOM ----------------
def create_master(dfs):
    df_all = pd.concat(dfs, ignore_index=True)
    master = (
        df_all
        .dropna(subset=["Assembly Area", "No of Parts", "Component name"])
        .drop_duplicates(subset=["Assembly Area", "No of Parts", "Component name"])
        .copy()
    )
    master["Sr No"] = master["No of Parts"].apply(
        lambda x: int(x) if pd.notna(x) else pd.NA
    )
    master = master.sort_values(
        by=["Sr No", "No of Parts", "Component name"]
    ).reset_index(drop=True)
    return master[["Sr No", "Assembly Area", "No of Parts", "Component name"]]


# ---------------- NORMALIZE ----------------
def normalize(df, master):
    merged = master.merge(
        df,
        on=["Assembly Area", "No of Parts", "Component name"],
        how="left"
    )
    merged = merged.fillna("NA")
    merged["Total Assembly weight"] = merged["Total Assembly weight"].replace("NA", pd.NA)
    merged["Total Assembly weight"] = merged.groupby("Assembly Area")[
        "Total Assembly weight"
    ].transform(lambda x: x.ffill().bfill())
    merged["Total Assembly weight"] = merged["Total Assembly weight"].fillna("NA")
    merged = merged.drop_duplicates(
        subset=["Assembly Area", "No of Parts", "Component name"]
    )
    merged = merged.sort_values(
        by=["Sr No", "No of Parts", "Component name"]
    ).reset_index(drop=True)
    return merged[["Sr No", "Assembly Area", "No of Parts", "Component name", *SPEC_COLS]]


# ---------------- FORMAT ----------------
def apply_format_memory(excel_buffer):
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb.active
    red_font = Font(color="FF0000")
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        non_key_cells = [
            cell for cell in row
            if headers[cell.column - 1] not in KEY_COLS
        ]
        all_na = all(cell.value == "NA" for cell in non_key_cells)
        if all_na:
            for cell in row:
                cell.font = red_font
        else:
            for cell in non_key_cells:
                if cell.value == "NA":
                    cell.font = red_font
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------- ANALYSIS HELPERS ----------------
def is_na(val):
    if pd.isna(val):
        return True
    val = str(val).strip().lower()
    return val in ["na", "nan", "", "none", "n/a", "-", "--"]


def get_presence_matrix(assembly_data, spec_cols=None):
    """
    spec_cols defaults to the washing-machine ANALYSIS_SPEC_COLS for backward
    compatibility, but callers should pass the job's stored
    analysis_spec_cols (JOB_STORE[job_id]["analysis_spec_cols"]) so AC jobs
    check presence against AC's actual spec columns (which don't include
    Thickness and do include Colour) instead of washing machine's.
    """
    cols_to_check = spec_cols or ANALYSIS_SPEC_COLS

    all_components = set()
    for df in assembly_data.values():
        all_components.update(df["Component name"].dropna().unique())

    matrix = pd.DataFrame(index=sorted(all_components))

    for company, df in assembly_data.items():
        temp = df.copy()
        temp["Component name"] = temp["Component name"].astype(str).str.strip()
        presence = []
        for comp in matrix.index:
            row = temp[temp["Component name"] == comp]
            if row.empty:
                presence.append(0)
                continue
            row = row.iloc[0]
            all_na = True
            for col in cols_to_check:
                val = row.get(col, None)
                if not is_na(val):
                    all_na = False
                    break
            presence.append(0 if all_na else 1)
        matrix[company] = presence

    return matrix


def get_top_weighted_components(assembly_data, top_n=5):
    """
    Finds the components within an assembly that are heaviest AND present
    (non-NA weight) across every company — i.e. fair cross-brand comparison
    points, not a component only one company happens to report.
 
    assembly_data: dict of {company: df}, already filtered to one assembly
                   (the output of get_assembly_data()).
    top_n: how many components to return (default 5).
 
    Returns a list of dicts, sorted heaviest-first:
        [{"component": "clutch plate", "avg_weight": 86.5,
          "weights_by_company": {"IFB": 90, "Whirlpool": 83, ...}}, ...]
    Returns an empty list if no component has valid weight data in every
    company (e.g. too few companies uploaded, or weight wasn't recorded).
    """
    companies = list(assembly_data.keys())
    if not companies:
        return []
    # collect every component name that appears anywhere in this assembly
    all_components = set()
    for df in assembly_data.values():
        all_components.update(df["Component name"].dropna().unique())
    candidates = []
    for comp in all_components:
        weights_by_company = {}
        all_present = True
 
        for company in companies:
            df = assembly_data[company]
            match = df[df["Component name"].astype(str).str.strip() == comp]
 
            if match.empty:
                all_present = False
                break
 
            weight_val = match.iloc[0].get("Weight (Grams/piece)", None)
            if is_na(weight_val):
                all_present = False
                break
 
            try:
                weights_by_company[company] = float(weight_val)
            except (TypeError, ValueError):
                all_present = False
                break
  
        if all_present and weights_by_company:
            avg_weight = sum(weights_by_company.values()) / len(weights_by_company)
            candidates.append({
                "component": comp,
                "avg_weight": round(avg_weight, 2),
                "weights_by_company": weights_by_company,
            })
 
    candidates.sort(key=lambda c: c["avg_weight"], reverse=True)
    return candidates[:top_n]


def get_assembly_data(data, assembly_name):
    result = {}
    for company, df in data.items():
        temp = df[df["Assembly Area"].str.lower().str.contains(assembly_name.lower(), na=False)]
        result[company] = temp.reset_index(drop=True)
    return result


def filter_by_unit(data, unit):
    """
    Scopes each company's dataframe to a single AC unit ("IDU" or "ODU").
    Washing machine dataframes have no "Unit" column at all, so this is a
    safe no-op for them — only AC jobs (which tag every row with its
    source unit in ac_bom_services.process_ac_files_api) are affected.
    """
    if not unit:
        return data
    result = {}
    for company, df in data.items():
        if "Unit" in df.columns:
            result[company] = df[df["Unit"].str.upper() == unit.upper()].reset_index(drop=True)
        else:
            result[company] = df
    return result


def compare_component_specs(data, component_name, spec_cols=None):
    """
    spec_cols defaults to washing-machine ANALYSIS_SPEC_COLS for backward
    compatibility — pass the job's stored analysis_spec_cols for AC jobs.
    """
    cols_to_use = spec_cols or ANALYSIS_SPEC_COLS

    rows = []
    for company, df in data.items():
        temp = df.copy()
        temp["Component name"] = temp["Component name"].astype(str).str.lower().str.strip()
        match = temp[temp["Component name"] == component_name.lower()]
        if match.empty:
            row_data = {"Company": company, "Component": component_name}
            for col in cols_to_use:
                row_data[col] = "NA"
        else:
            row = match.iloc[0]
            row_data = {"Company": company, "Component": component_name}
            for col in cols_to_use:
                row_data[col] = row.get(col, "NA")
        rows.append(row_data)
    return pd.DataFrame(rows)


def highlight_differences(df):
    insights = []
    cols = df.columns.drop(["Company", "Component"])
    for col in cols:
        values = df[col].astype(str).str.lower().unique()
        if len(values) > 1:
            insights.append(f"{col} differs across companies: {list(values)}")
    return insights


def best_component_analysis(df):
    insights = []

    def to_float(x):
        try:
            return float(str(x).replace("mm", "").strip())
        except:
            return None

    if "Thickness (mm)" in df.columns:
        temp = df.copy()
        temp["Thickness (mm)"] = temp["Thickness (mm)"].apply(to_float)
        temp = temp.dropna(subset=["Thickness (mm)"])
        if not temp.empty:
            best = temp.loc[temp["Thickness (mm)"].idxmax()]
            insights.append(f"Best strength (thickness): {best['Company']} ({best['Thickness (mm)']} mm)")

    if "Weight (Grams/piece)" in df.columns:
        temp = df.copy()
        temp["Weight (Grams/piece)"] = pd.to_numeric(temp["Weight (Grams/piece)"], errors="coerce")
        temp = temp.dropna(subset=["Weight (Grams/piece)"])
        if not temp.empty:
            best = temp.loc[temp["Weight (Grams/piece)"].idxmin()]
            insights.append(f"Lightest component: {best['Company']} ({best['Weight (Grams/piece)']} g)")

    if "Material" in df.columns:
        materials = df[["Company", "Material"]].dropna()
        if len(materials["Material"].unique()) > 1:
            insights.append("Material differs across companies → depends on quality")

    # AC-specific spec column — washing machine BOMs don't have this, so it
    # naturally no-ops there since the column won't exist.
    if "Colour" in df.columns:
        colours = df[["Company", "Colour"]].dropna()
        if len(colours["Colour"].unique()) > 1:
            insights.append("Colour differs across companies")

    if "Manufacturing Process" in df.columns:
        procs = df["Manufacturing Process"].dropna().unique()
        if len(procs) > 1:
            insights.append(f"Different manufacturing processes used: {list(procs)}")

    if "Characteristic (eg-special point)" in df.columns:
        chars = df[["Company", "Characteristic (eg-special point)"]].dropna()
        unique_chars = chars["Characteristic (eg-special point)"].unique()
        if len(unique_chars) > 1:
            insights.append("Different characteristics across companies:")
        for _, row in chars.iterrows():
            insights.append(f"  {row['Company']}: {row['Characteristic (eg-special point)']}")

    return insights


def analyze_multiple_components(multi_df):
    all_insights = []
    for comp, df in multi_df.groupby("Component"):
        all_insights.append(f"=== {comp.upper()} ===")
        for i in best_component_analysis(df):
            all_insights.append(i)
    return all_insights


# ---------------- MAIN PIPELINE ----------------
def process_files_api(file_bytes_dict, company_names):
    """
    file_bytes_dict: {original_filename: bytes}
    company_names: [list of company names in same order as files]
    """
    job_id = str(uuid.uuid4())
    dfs = {}

    for (name, content), company in zip(file_bytes_dict.items(), company_names):
        df = load_bm_sheet(io.BytesIO(content))
        df = clean_df(df)
        dfs[company] = df  # key is now clean company name

    master = create_master(list(dfs.values()))

    normalized = {}
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for company, df in dfs.items():
            norm_df = normalize(df, master)
            normalized[company] = norm_df  # store by company name

            excel_buffer = io.BytesIO()
            norm_df.to_excel(excel_buffer, index=False)
            excel_buffer.seek(0)
            excel_buffer = apply_format_memory(excel_buffer)

            zipf.writestr(f"{company}_normalized.xlsx", excel_buffer.getvalue())

    JOB_STORE[job_id] = {
        "normalized": normalized,
        "product_type": "washing_machine",
        "analysis_spec_cols": ANALYSIS_SPEC_COLS,
    }

    zip_buffer.seek(0)
    return job_id, zip_buffer