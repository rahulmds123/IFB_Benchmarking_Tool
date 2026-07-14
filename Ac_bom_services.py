"""
AC (split-unit) BOM cleaning/normalization pipeline — parallel to
bom_services.py's washing machine pipeline, but AC BOM files have two
sheets per file (IDU = indoor unit, ODU = outdoor unit) instead of one,
and a different spec column set (Colour instead of Thickness).

process_ac_files_api() combines each company's IDU + ODU normalized data
into a single per-company dataframe before storing it in JOB_STORE — IDU
and ODU components just become part of the same "Assembly Area" universe
(same pattern washing machine assemblies already use), so every other
endpoint (presence matrix, top components, spec comparison, PDF report)
works on AC jobs without any AC-specific branching. The only thing that
differs per job is which spec columns to look at, which is why
bom_services.py's get_presence_matrix/compare_component_specs now take an
explicit spec_cols argument instead of assuming a single hardcoded set.
"""

import pandas as pd
import uuid
import zipfile
import io
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.services.bom_services import JOB_STORE

# --------------------------------------------------------
# Assembly aliases (add AC-specific ones here if needed)
# --------------------------------------------------------
ASSEMBLY_ALIASES = {}

# --------------------------------------------------------
# Column Mapping
# --------------------------------------------------------
AC_IDU_COLUMN_MAP = {
    "Basic Information (According to BOM)": "Sr No",
    "Component name/Child part": "Component name",
}

AC_ODU_COLUMN_MAP = {
    "Basic Information (According to BOM)": "Sr No",
    "Component name/Child part": "Component name",
    # ODU source file has a mixed-case typo in weight column
    "Weight (GraMS/piece)": "Weight (Grams/piece)",
}

# --------------------------------------------------------
# Spec Columns
# --------------------------------------------------------
AC_SPEC_COLS = [
    "Manufacturing Process",
    "Position (Assembled where)",
    "Dimensions/Specs(mm)",
    "Material",
    "Colour",
    "Weight (Grams/piece)",
    "Number of Part",
    "Total Weight",
    "Total Assembly weight",
    "Characteristic (eg-special point)",
]

AC_ANALYSIS_SPEC_COLS = [
    "Manufacturing Process",
    "Position (Assembled where)",
    "Dimensions/Specs(mm)",
    "Material",
    "Colour",
    "Weight (Grams/piece)",
    "Characteristic (eg-special point)",
]

KEY_COLS = {"Sr No", "Assembly Area", "No of Parts", "Component name"}


# --------------------------------------------------------
# LOAD
# --------------------------------------------------------
def load_ac_sheets(file):
    xls = pd.ExcelFile(file)

    idu_sheet = next(
        (s for s in xls.sheet_names if "Detailed BM Chart - IDU" in s), None
    )
    odu_sheet = next(
        (s for s in xls.sheet_names if "Detailed BM Chart - ODU" in s), None
    )

    if idu_sheet is None:
        raise ValueError("IDU BM Chart not found")
    if odu_sheet is None:
        raise ValueError("ODU BM Chart not found")

    df_idu = pd.read_excel(xls, sheet_name=idu_sheet)
    df_odu = pd.read_excel(xls, sheet_name=odu_sheet)

    return df_idu, df_odu


# --------------------------------------------------------
# CLEAN
# --------------------------------------------------------
def clean_ac_df(df, column_map):

    # 1. Normalize column names
    df.columns = (
        df.columns
        .str.replace("\n", " ", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    # 2. Rename known columns to canonical names
    df = df.rename(columns=column_map)
    df = df.rename(columns={"Dimensions/Specs (mm)": "Dimensions/Specs(mm)"})

    # ----------------------------------------------------
    # ODU Fix
    # Some ODU files use "compressor" instead of
    # "Assembly Area" as the column header.
    # ----------------------------------------------------
    if "Assembly Area" not in df.columns:
        for col in df.columns:
            col_clean = str(col).strip().lower()
            if col_clean == "compressor":
                df.rename(columns={col: "Assembly Area"}, inplace=True)
                break

    # Fallback:
    # Whatever the second column is, treat it as Assembly Area
    if "Assembly Area" not in df.columns:
        cols = list(df.columns)
        if len(cols) > 1:
            cols[1] = "Assembly Area"
            df.columns = cols

    # 3. Save the sub-header row (row index 1 = "W-Width", "D-Depth" etc.)
    #    BEFORE dropping it — we need it to detect which Unnamed col is which
    #    Note: row 0 is NaN for IDU, row 0 has real sub-headers for ODU
    #    So we check which of the first two rows has the dimension labels
    header_row = None
    for i in [0, 1]:
        row = df.iloc[i]
        row_str = " ".join(str(v) for v in row.values).lower()
        if "width" in row_str or "depth" in row_str or "height" in row_str:
            header_row = row.copy()
            break

    # drop both junk rows (row 0 and row 1), real data starts at row 2
    df = df.iloc[2:].reset_index(drop=True)

    # 4. Detect and rename dimension sub-columns using saved header row
    if header_row is not None:
        rename_dict = {}
        for col in df.columns:
            header = str(header_row.get(col, "")).strip().lower()
            if "width" in header:
                rename_dict[col] = "Dim_Width"
            elif "height" in header:
                rename_dict[col] = "Dim_Height"
            elif "depth" in header:
                rename_dict[col] = "Dim_Depth"
            elif "diameter" in header:
                rename_dict[col] = "Dim_Diameter"
            elif "thickness" in header:
                rename_dict[col] = "Dim_Thickness"
        df = df.rename(columns=rename_dict)

    # 5. Combine dimension sub-columns into one Dimensions/Specs(mm) string
    #    Format: "W=880, D=311, H=24.8, T=0.75"
    #    Overwrites the parent column (which may have a partial value anyway)
    dim_cols = {
        "Dim_Width": "W",
        "Dim_Height": "H",
        "Dim_Depth": "D",
        "Dim_Diameter": "Dia",
        "Dim_Thickness": "T",
    }
    existing_dim_cols = {k: v for k, v in dim_cols.items() if k in df.columns}

    def combine_dimensions(row):
        parts = []
        for col, short in existing_dim_cols.items():
            val = row.get(col, None)
            if pd.notna(val) and str(val).strip() not in ("", "nan", "NA"):
                parts.append(f"{short}={val}")
        return ", ".join(parts) if parts else pd.NA

    df["Dimensions/Specs(mm)"] = df.apply(combine_dimensions, axis=1)

    # 6. Drop helper/unnamed columns now that dims are combined
    drop_cols = (
        list(existing_dim_cols.keys())
        + ["Images"]
        + [c for c in df.columns if c.startswith("Unnamed")]
    )
    df = df.drop(columns=drop_cols, errors="ignore")

    # 7. Normalize all cell values (same as washing machine clean_df)
    for col in df.select_dtypes(include="object").columns:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace("\n", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    df = df.replace("nan", pd.NA).replace("NA", pd.NA)

    # 8. Normalize key columns
    if "Assembly Area" not in df.columns:
        raise ValueError(
            f"Assembly Area column not found.\nColumns are:\n{list(df.columns)}"
        )
    df["Assembly Area"] = (
        df["Assembly Area"]
        .astype(str).str.strip().str.lower()
        .str.replace(r"\s+", " ", regex=True)
        .replace("nan", pd.NA)
        .replace(ASSEMBLY_ALIASES)
    )
    df["Component name"] = (
        df["Component name"]
        .astype(str).str.strip().str.lower()
        .replace("nan", pd.NA)
    )
    df["No of Parts"] = pd.to_numeric(df["No of Parts"], errors="coerce").round(2)
    df["Sr No"] = pd.to_numeric(df["Sr No"], errors="coerce")

    # drop rows with no valid No of Parts (leftover junk rows)
    df = df.dropna(subset=["No of Parts"]).copy()

    # forward-fill Assembly Area
    df["Assembly Area"] = df["Assembly Area"].ffill()

    return df


# --------------------------------------------------------
# MASTER BOM — same logic as washing machine
# --------------------------------------------------------
def create_ac_master(dfs):
    df_all = pd.concat(dfs, ignore_index=True)

    master = (
        df_all
        .dropna(subset=["Assembly Area", "No of Parts", "Component name"])
        .drop_duplicates(subset=["Assembly Area", "No of Parts", "Component name"])
        .copy()
    )

    # Sr No = integer prefix of No of Parts
    master["Sr No"] = master["No of Parts"].apply(lambda x: int(float(x)))

    master = master.sort_values(
        by=["Sr No", "No of Parts", "Component name"]
    ).reset_index(drop=True)

    return master[["Sr No", "Assembly Area", "No of Parts", "Component name"]]


# --------------------------------------------------------
# NORMALIZE — same left-join pattern as washing machine
# --------------------------------------------------------
def normalize_ac(df, master):
    """
    Left-joins company df against master BOM, fills missing specs with NA,
    propagates Total Assembly weight within each assembly, dedupes and sorts.
    Only uses AC_SPEC_COLS that actually exist in this df.
    """
    available_spec_cols = [c for c in AC_SPEC_COLS if c in df.columns]

    merged = master.merge(
        df,
        on=["Assembly Area", "No of Parts", "Component name"],
        how="left",
        suffixes=("", "_comp"),
    )
    merged = merged.fillna("NA")

    if "Total Assembly weight" in merged.columns:
        merged["Total Assembly weight"] = merged["Total Assembly weight"].replace("NA", pd.NA)
        merged["Total Assembly weight"] = merged.groupby("Assembly Area")[
            "Total Assembly weight"
        ].transform(lambda x: x.ffill().bfill())
        merged["Total Assembly weight"] = merged["Total Assembly weight"].fillna("NA")

    merged = merged.drop_duplicates(
        subset=["Assembly Area", "No of Parts", "Component name"]
    )

    merged = merged.sort_values(
        by=["Sr No", "No of Parts", "Component name"]
    ).reset_index(drop=True)

    return merged[
        ["Sr No", "Assembly Area", "No of Parts", "Component name", *available_spec_cols]
    ]


# --------------------------------------------------------
# FORMAT — same red-font logic as washing machine
# --------------------------------------------------------
def apply_format_memory_ac(excel_buffer):
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb.active
    red_font = Font(color="FF0000")
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        non_key_cells = [
            cell for cell in row
            if headers[cell.column - 1] not in KEY_COLS
        ]
        all_na = all(cell.value == "NA" for cell in non_key_cells)
        if all_na:
            for cell in row:
                cell.font = red_font
        else:
            for cell in non_key_cells:
                if cell.value == "NA":
                    cell.font = red_font
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --------------------------------------------------------
# MAIN PIPELINE — mirrors bom_services.process_files_api's shape so it
# plugs into the same JOB_STORE and every downstream endpoint (assemblies,
# presence-matrix, top-components, multi-component, report) works
# unmodified. IDU + ODU get combined into one per-company dataframe: their
# components just live under different "Assembly Area" values, the same
# way washing machine assemblies (e.g. "drum assembly", "motor assembly")
# already coexist within one company dataframe.
# --------------------------------------------------------
def process_ac_files_api(file_bytes_dict, company_names):
    """
    file_bytes_dict: {original_filename: bytes} — each file must contain
                      both a "Detailed BM Chart - IDU..." and a
                      "Detailed BM Chart - ODU..." sheet.
    company_names: [list of company names in same order as files]

    Returns (job_id, zip_buffer) — same shape as process_files_api, so
    main.py's /upload endpoint can call either one interchangeably based
    on the product_type the user selected at upload.
    """
    job_id = str(uuid.uuid4())

    idu_by_company = {}
    odu_by_company = {}

    for (name, content), company in zip(file_bytes_dict.items(), company_names):
        idu_raw, odu_raw = load_ac_sheets(io.BytesIO(content))
        idu_by_company[company] = clean_ac_df(idu_raw, AC_IDU_COLUMN_MAP)
        odu_by_company[company] = clean_ac_df(odu_raw, AC_ODU_COLUMN_MAP)

    master_idu = create_ac_master(list(idu_by_company.values()))
    master_odu = create_ac_master(list(odu_by_company.values()))

    normalized = {}
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for company in company_names:
            idu_norm = normalize_ac(idu_by_company[company], master_idu)
            odu_norm = normalize_ac(odu_by_company[company], master_odu)

            # tag each row with its source unit before combining — this is
            # what lets the frontend and API scope everything (assemblies,
            # presence matrix, components, analysis, reports) to just IDU
            # or just ODU instead of a flat merged list.
            idu_norm = idu_norm.copy()
            odu_norm = odu_norm.copy()
            idu_norm["Unit"] = "IDU"
            odu_norm["Unit"] = "ODU"

            combined = pd.concat([idu_norm, odu_norm], ignore_index=True, sort=False)
            combined = combined.fillna("NA")
            normalized[company] = combined

            for label, norm_df in (("IDU", idu_norm), ("ODU", odu_norm)):
                excel_buffer = io.BytesIO()
                norm_df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)
                excel_buffer = apply_format_memory_ac(excel_buffer)
                zipf.writestr(f"{company}_{label}_normalized.xlsx", excel_buffer.getvalue())

    JOB_STORE[job_id] = {
        "normalized": normalized,
        "product_type": "ac",
        "analysis_spec_cols": AC_ANALYSIS_SPEC_COLS,
    }

    zip_buffer.seek(0)
    return job_id, zip_buffer